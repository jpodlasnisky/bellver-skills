#!/usr/bin/env python3
import re
import sys
from pathlib import Path


def extrair_cnpjs_de_textos(textos: list[str]) -> list[str]:
    cnpjs: list[str] = []
    for texto in textos:
        for candidato in re.findall(r"[\d./-]{14,20}", texto):
            digitos = re.sub(r"\D", "", candidato)
            if len(digitos) == 14 and digitos not in cnpjs:
                cnpjs.append(digitos)
    return cnpjs


def ler_txt_ou_csv(caminho: Path) -> list[str]:
    texto = caminho.read_text(encoding="utf-8-sig", errors="ignore")
    return extrair_cnpjs_de_textos([texto])


def ler_xlsx(caminho: Path) -> list[str]:
    from openpyxl import load_workbook

    wb = load_workbook(caminho, data_only=True, read_only=True)
    textos = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for valor in row:
                if valor is not None:
                    textos.append(str(valor))
    return extrair_cnpjs_de_textos(textos)


def main():
    if len(sys.argv) != 2:
        print("uso: ler_lista_cnpjs.py <arquivo.txt|.csv|.xlsx>", file=sys.stderr)
        sys.exit(2)

    caminho = Path(sys.argv[1])
    if not caminho.exists():
        print(f"ERRO: arquivo não encontrado: {caminho}", file=sys.stderr)
        sys.exit(1)

    sufixo = caminho.suffix.lower()
    if sufixo in (".xlsx", ".xlsm"):
        cnpjs = ler_xlsx(caminho)
    elif sufixo in (".txt", ".csv"):
        cnpjs = ler_txt_ou_csv(caminho)
    else:
        print(f"ERRO: formato não suportado: {sufixo} (use .txt, .csv ou .xlsx)", file=sys.stderr)
        sys.exit(1)

    if not cnpjs:
        print("ERRO: nenhum CNPJ (14 dígitos) encontrado no arquivo.", file=sys.stderr)
        sys.exit(1)

    for cnpj in cnpjs:
        print(cnpj)


if __name__ == "__main__":
    main()
